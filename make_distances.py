#!/usr/bin/env python3

import json
from openpyxl import Workbook
import bisect
import os
import subprocess
import signal
import time
import requests

def city_name(city):
    return "%s, %s, %d" % (city["properties"]["name"], city["properties"]["countyMn"], city["properties"]["natCode"])

def is_big_city(city):
    rank = city["properties"]["rank"]
    return rank in ["0", "I", "II", "III"]

graphhopper_proc = None
try:
    graphhopper_proc = subprocess.Popen(["./graphhopper.sh"], stdin=None, stdout=subprocess.DEVNULL, stderr=subprocess.DEVNULL, shell=False, preexec_fn=os.setsid)

    while True:
        try:
            requests.get("http://localhost:8989/health")
            break
        except Exception as e:
            print(e)
            time.sleep(1)

    with open(os.getenv("DISTANTE_INPUT") or "ro_localitati_punct.geojson", "r") as file:
        data = json.load(file)

    cities = data["features"]
    cities_max = os.getenv("DISTANTE_MAX")
    if cities_max and cities_max.isdecimal():
        cities = cities[0:int(cities_max)]

    cities.sort(key=lambda city: city["properties"]["countyMn"])

    #distances = np.zeros((len(cities), len(cities)))

    counties = set(map(lambda city: city["properties"]["countyMn"], cities))
    for county in counties:
        wb = Workbook()
        ws = wb.active

        city_start = bisect.bisect_left(cities, county, key=lambda city: city["properties"]["countyMn"])
        city_end = bisect.bisect_right(cities, county, key=lambda city: city["properties"]["countyMn"])
        print("Judetul %s: %d-%d" % (county, city_start, city_end))
        for i in range(city_start + 1, city_end):
            print("%s: %d/%d" % (county, i, city_end - 1))
            for j in range(city_start, i):
                if not is_big_city(cities[i]) and not is_big_city(cities[j]):
                    continue
                r = requests.post("http://localhost:8989/route", json={
                    "points": [cities[i]["geometry"]["coordinates"] for i in [i, j]],
                    "profile": "car",
                    "instructions": False,
                    "calc_points": False,
                })
                result = r.json()

                if "paths" not in result or len(result["paths"]) == 0:
                    print("Drum izolat: %s - %s" % (city_name(cities[i]), city_name(cities[j])))
                    continue
                distance = sorted(result["paths"], key=lambda x: x["distance"])[0]["distance"] / 1000
                ws.cell(row=i+2-city_start, column=j+2-city_start).value = distance
                ws.cell(row=j+2-city_start, column=i+2-city_start).value = distance

        for i in range(city_start, city_end):
            ws.cell(row=1, column=i+2-city_start).value = city_name(cities[i])
            ws.cell(row=i+2-city_start, column=1).value = city_name(cities[i])

        wb.save(f"distante_{county}.xlsx")

    #with open(os.getenv("DISTANTE_OUTPUT") or "distante.csv", "w") as csv_file:
    #    csv_writer = csv.writer(csv_file)
    #    csv_writer.writerow([""] + [city_name(city) for city in cities])
    #    for i in range(len(cities)):
    #        csv_writer.writerow([city_name(cities[i])] + ["%.1f" % d for d in distances[i]])
finally:
    if graphhopper_proc:
        os.killpg(os.getpgid(graphhopper_proc.pid), signal.SIGTERM)
